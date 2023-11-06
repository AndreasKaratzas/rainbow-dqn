"""Utility functions for the project.
"""

import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from pathlib import Path
import matplotlib.pyplot as plt


def create_data_dir(export_data_path):
    """Create a directory for saving the model, memory, and log.

    Parameters
    ----------
    export_data_path : pathlib.Path
        Path to the directory where the data will be saved.

    Returns
    -------
    model_save_dir : pathlib.Path
        Path to the directory where the model will be saved.
    memory_save_dir : pathlib.Path
        Path to the directory where the memory will be saved.
    log_save_dir : pathlib.Path
        Path to the directory where the log will be saved.
    """
    datetime_tag = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    model_save_dir = export_data_path / datetime_tag / "model"
    model_save_dir.mkdir(parents=True, exist_ok=True)
    memory_save_dir = export_data_path / datetime_tag / "memory"
    memory_save_dir.mkdir(parents=True, exist_ok=True)
    log_save_dir = export_data_path / datetime_tag / "log"
    log_save_dir.mkdir(parents=True, exist_ok=True)
    return model_save_dir, memory_save_dir, log_save_dir, datetime_tag


def get_chkpt(resume: str = None):
    """Get the latest checkpoint data.

    Parameters
    ----------
    resume : str
        Path to the parent directory where the checkpoint data is saved.

    Returns
    -------
    model_checkpoint : str
        Path to the latest model checkpoint.
    mem_checkpoint : str
        Path to the latest memory checkpoint.
    """
    if resume:
        model_checkpoint = Path(resume) / "model"
        mem_checkpoint = Path(resume) / "memory"
    else:
        model_checkpoint = None
        mem_checkpoint = None
    return model_checkpoint, mem_checkpoint


def stats_printer(
    d: dict = {}, 
    indent: int = 0
):
    """Print the statistics of the agent.
    
    Parameters
    ----------
    d : dict
        Dictionary of statistics.
    indent : int
        Indentation level.
    """
    np.set_printoptions(precision=4, suppress=True, formatter={'float': '{: 0.3f}'.format})
    for key, value in d.items():
        print('\t' * indent + str(key))
        if isinstance(value, dict):
            stats_printer(value, indent+1)
        else:
            print('\t' * (indent+1) + str(value))


def test_stats(towards_target: list = [], mission_status: list = []):
    """Gives important percentages regarding the model's test process.
    """

    labels_1 = ["Converged", "Diverged"]
    labels_2 = ["Success", "Failed"]

    y_1 = np.array([sum(towards_target), len(towards_target) - sum(towards_target)], dtype=np.int32)
    y_2 = np.array([sum(mission_status), len(mission_status) - sum(mission_status)], dtype=np.int32)

    fig, (ax_1, ax_2) = plt.subplots(2, 1)
    fig.suptitle('Test stats')

    ax_1.pie(y_1, labels=labels_1, autopct='%1.1f%%')
    ax_1.title.set_text('Action correctness w.r.t Target')

    ax_2.pie(y_2, labels=labels_2, autopct='%1.1f%%')
    ax_2.title.set_text('Mission success rate')
    
    plt.tight_layout()
    plt.show()


def save_3d_array_to_xlsx(arr, filename):
    if arr.shape != (4, 64, 64):
        raise ValueError("Input array must have shape (4, 64, 64)")

    wb = Workbook()
    
    # Remove the default sheet created by Workbook()
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)
    
    for i in range(arr.shape[0]):
        sheet_name = f'Sheet{i+1}'
        ws = wb.create_sheet(sheet_name)
        
        for j in range(arr.shape[1]):
            for k in range(arr.shape[2]):
                ws.cell(row=j+1, column=k+1, value=arr[i, j, k])
    
    wb.save(filename)


def experiment_data_plots(export_data_dir, datetime_tag):
    """Plots training loss progress and mean accumulated episode reward.
    """
    filename = export_data_dir / datetime_tag / "log" / "log"
    df = pd.read_csv(filename, delim_whitespace=True)
    plt.subplot(2, 1, 1)
    plt.title('Training stats')
    plt.plot(df.Step, df.MeanLoss)
    plt.ylabel('Mean Loss')
    plt.subplot(2, 1, 2)
    plt.plot(df.Step, df.MeanReward)
    plt.xlabel('Step')
    plt.ylabel('Mean Reward')
    plt.tight_layout()
    plt.show()